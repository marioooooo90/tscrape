import os
import sqlite3
import json
import csv
import asyncio
from telethon import TelegramClient
from telethon.tl.types import MessageMediaPhoto, MessageMediaDocument, User, PeerChannel, MessageMediaPoll
from telethon.errors import FloodWaitError, RPCError, SessionPasswordNeededError, PhoneNumberInvalidError
import aiohttp
import sys
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from io import BytesIO
import math
import pickle
from telethon.sessions import StringSession
from datetime import datetime, timezone
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
import base64
from tqdm import tqdm
import time
import psutil

# Add this near the top of the file, after the imports
continuous_scraping_active = False
continuous_scraping_task = None
scraping_log = []

def display_ascii_art():
    WHITE = "\033[97m"
    RESET = "\033[0m"
    
    art = r"""
___________________  _________
\__    ___/  _____/ /   _____/
  |    | /   \  ___ \_____  \ 
  |    | \    \_\  \/        \
  |____|  \______  /_______  /
                 \/        \/
    """
    
    print(WHITE + art + RESET)

display_ascii_art()

STATE_FILE = 'state.json'

# Constants for storage
CREDENTIALS_FILE = 'credentials.pkl'
CHANNELS_FILE = 'channels.pkl'

# Add these constants at the top with other constants
GOOGLE_CREDS_FILE = 'google_token.pickle'
SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive.file',
    'https://www.googleapis.com/auth/drive'  # Add full drive access
]

def save_credentials(api_id, api_hash, phone, session_string=None):
    with open(CREDENTIALS_FILE, 'wb') as f:
        pickle.dump({
            'api_id': api_id,
            'api_hash': api_hash,
            'phone': phone,
            'session_string': session_string
        }, f)

def load_credentials():
    try:
        with open(CREDENTIALS_FILE, 'rb') as f:
            return pickle.load(f)
    except:
        return None

def save_channels(channels):
    with open(CHANNELS_FILE, 'wb') as f:
        pickle.dump(channels, f)

def load_channels():
    try:
        with open(CHANNELS_FILE, 'rb') as f:
            return pickle.load(f)
    except:
        return {}

def initialize_state():
    credentials = load_credentials()
    if credentials:
        return {
            'api_id': credentials['api_id'],
            'api_hash': credentials['api_hash'],
            'phone': credentials['phone'],
            'channels': load_channels(),
            'scrape_media': True,
        }
    return {
        'api_id': None,
        'api_hash': None,
        'phone': None,
        'channels': load_channels(),
        'scrape_media': True,
    }

state = initialize_state()

if not state['api_id'] or not state['api_hash'] or not state['phone']:
    state['api_id'] = int(input("Enter your API ID: "))
    state['api_hash'] = input("Enter your API Hash: ")
    state['phone'] = input("Enter your phone number: ")
    save_credentials(state['api_id'], state['api_hash'], state['phone'])

client = TelegramClient('session', state['api_id'], state['api_hash'])

async def get_channel_name(channel_id):
    try:
        if channel_id.startswith('-'):
            entity = await client.get_entity(PeerChannel(int(channel_id)))
        else:
            entity = await client.get_entity(channel_id)
        # Clean the channel name to be filesystem-friendly
        channel_name = "".join(c for c in entity.title if c.isalnum() or c in (' ', '-', '_')).strip()
        return f"{channel_name}_{channel_id}"
    except Exception as e:
        print(f"Error getting channel name: {str(e)}")
        return channel_id

def save_message_to_db(channel, message, sender):
    channel_dir = os.path.join(os.getcwd(), state['channels'][channel]['folder_name'])
    os.makedirs(channel_dir, exist_ok=True)

    db_file = os.path.join(channel_dir, f'{channel}.db')
    conn = sqlite3.connect(db_file)
    c = conn.cursor()
    
    # Check if table exists and its schema
    c.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='messages'")
    table_exists = c.fetchone()
    
    # If table exists with old schema, drop it
    if table_exists and ('reactions' not in table_exists[0] or 'views' not in table_exists[0]):
        c.execute('DROP TABLE messages')
        conn.commit()
    
    # Create table with new schema including all engagement metrics
    c.execute(f'''CREATE TABLE IF NOT EXISTS messages
                  (id INTEGER PRIMARY KEY, message_id INTEGER, date TEXT, sender_id INTEGER, 
                   first_name TEXT, last_name TEXT, username TEXT, message TEXT, 
                   formatted_message TEXT, entities TEXT, media_type TEXT, media_path TEXT, 
                   reply_to INTEGER, views INTEGER, forwards INTEGER, replies INTEGER,
                   reactions TEXT, post_author TEXT, is_pinned INTEGER)''')

    # Convert message entities to JSON-serializable format
    entities_data = None
    if hasattr(message, 'entities') and message.entities:
        entities_data = [{
            'type': str(entity.__class__.__name__),
            'offset': entity.offset,
            'length': entity.length,
            'url': getattr(entity, 'url', None)
        } for entity in message.entities]

    # Get reactions data
    reactions_data = None
    if hasattr(message, 'reactions') and message.reactions:
        reactions_data = {
            'total_count': getattr(message.reactions, 'total_count', 0),
            'reactions': [{
                'emoticon': getattr(reaction, 'emoticon', None),
                'count': reaction.count,
                'reaction_type': str(reaction.reaction),
            } for reaction in message.reactions.results]
        }

    try:
        c.execute('''INSERT OR REPLACE INTO messages 
                     (message_id, date, sender_id, first_name, last_name, username, 
                      message, formatted_message, entities, media_type, media_path, reply_to,
                      views, forwards, replies, reactions, post_author, is_pinned)
                     VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                  (message.id, 
                   message.date.strftime('%Y-%m-%d %H:%M:%S'), 
                   message.sender_id,
                   getattr(sender, 'first_name', None) if isinstance(sender, User) else None, 
                   getattr(sender, 'last_name', None) if isinstance(sender, User) else None,
                   getattr(sender, 'username', None) if isinstance(sender, User) else None,
                   message.message,
                   getattr(message, 'text', message.message),  # Preserve original formatting
                   json.dumps(entities_data) if entities_data else None,
                   message.media.__class__.__name__ if message.media else None, 
                   None,
                   message.reply_to_msg_id if hasattr(message, 'reply_to') else None,
                   getattr(message, 'views', 0),  # Views count
                   getattr(message, 'forwards', 0),  # Forwards count
                   getattr(message, 'replies', 0) if hasattr(message, 'replies') else 0,  # Replies count
                   json.dumps(reactions_data) if reactions_data else None,  # Reactions data
                   getattr(message, 'post_author', None),  # Post author (for channels)
                   1 if getattr(message, 'pinned', False) else 0))  # Is pinned message
        conn.commit()
    except Exception as e:
        print(f"Error saving message {message.id}: {str(e)}")
    finally:
        conn.close()

MAX_RETRIES = 5

async def download_media(channel, message):
    if not message.media or not state['scrape_media']:
        return None

    channel_dir = os.path.join(os.getcwd(), state['channels'][channel]['folder_name'])
    media_folder = os.path.join(channel_dir, 'media')
    os.makedirs(media_folder, exist_ok=True)    
    media_file_name = None
    if isinstance(message.media, MessageMediaPhoto):
        media_file_name = message.file.name or f"{message.id}.jpg"
    elif isinstance(message.media, MessageMediaDocument):
        media_file_name = message.file.name or f"{message.id}.{message.file.ext if message.file.ext else 'bin'}"
    
    if not media_file_name:
        print(f"Unable to determine file name for message {message.id}. Skipping download.")
        return None
    
    media_path = os.path.join(media_folder, media_file_name)
    
    if os.path.exists(media_path):
        print(f"Media file already exists: {media_path}")
        return media_path

    retries = 0
    while retries < MAX_RETRIES:
        try:
            if isinstance(message.media, MessageMediaPhoto):
                media_path = await message.download_media(file=media_folder)
            elif isinstance(message.media, MessageMediaDocument):
                media_path = await message.download_media(file=media_folder)
            if media_path:
                print(f"Successfully downloaded media to: {media_path}")
            break
        except (TimeoutError, aiohttp.ClientError, RPCError) as e:
            retries += 1
            print(f"Retrying download for message {message.id}. Attempt {retries}...")
            await asyncio.sleep(2 ** retries)
    return media_path

async def rescrape_media(channel):
    channel_dir = os.path.join(os.getcwd(), channel)
    db_file = os.path.join(channel_dir, f'{channel}.db')
    conn = sqlite3.connect(db_file)
    c = conn.cursor()
    c.execute('SELECT message_id FROM messages WHERE media_type IS NOT NULL AND media_path IS NULL')
    rows = c.fetchall()
    conn.close()

    total_messages = len(rows)
    if total_messages == 0:
        print(f"No media files to reprocess for channel {channel}.")
        return

    for index, (message_id,) in enumerate(rows):
        try:
            entity = await client.get_entity(PeerChannel(int(channel)))
            message = await client.get_messages(entity, ids=message_id)
            media_path = await download_media(channel, message)
            if media_path:
                conn = sqlite3.connect(db_file)
                c = conn.cursor()
                c.execute('''UPDATE messages SET media_path = ? WHERE message_id = ?''', (media_path, message_id))
                conn.commit()
                conn.close()
            
            progress = (index + 1) / total_messages * 100
            sys.stdout.write(f"\rReprocessing media for channel {channel}: {progress:.2f}% complete")
            sys.stdout.flush()
        except Exception as e:
            print(f"Error reprocessing message {message_id}: {e}")
    print()

async def scrape_channel(channel, offset_id, force_rescrape=False):
    try:
        if channel not in state['channels'] or 'folder_name' not in state['channels'][channel]:
            print("\nAnalyzing channel...")
            folder_name = await get_channel_name(channel)
            state['channels'][channel] = {
                'last_id': 0,
                'folder_name': folder_name
            }
            save_channels(state['channels'])

        # Reset offset_id if force_rescrape is True
        if force_rescrape:
            offset_id = 0
            # Clear existing database if it exists
            channel_dir = os.path.join(os.getcwd(), state['channels'][channel]['folder_name'])
            db_file = os.path.join(channel_dir, f'{channel}.db')
            if os.path.exists(db_file):
                os.remove(db_file)
                print(f"\nResetting database for channel: {state['channels'][channel]['folder_name']}")

        if channel.startswith('-'):
            entity = await client.get_entity(PeerChannel(int(channel)))
        else:
            entity = await client.get_entity(channel)

        print("\nCounting messages in channel...")
        total_messages = 0
        message_count = 0
        last_update = 0
        
        async for _ in client.iter_messages(entity, offset_id=offset_id, reverse=True):
            total_messages += 1
            message_count += 1
            
            # Update progress every 100 messages or when 1 second has passed
            current_time = asyncio.get_event_loop().time()
            if message_count >= 100 or (current_time - last_update) >= 1:
                sys.stdout.write(f"\rCounting messages: {total_messages} found...")
                sys.stdout.flush()
                message_count = 0
                last_update = current_time
                
        print(f"\nFound {total_messages} messages to process")

        if total_messages == 0:
            print(f"No new messages found in channel {channel}.")
            return

        processed_messages = 0
        last_message_id = None

        print("\nStarting message scraping...")
        async for message in client.iter_messages(entity, offset_id=offset_id, reverse=True):
            try:
                sender = await message.get_sender()
                save_message_to_db(channel, message, sender)

                if state['scrape_media'] and message.media:
                    media_path = await download_media(channel, message)
                    if media_path:
                        conn = sqlite3.connect(os.path.join(os.getcwd(), 
                                                          state['channels'][channel]['folder_name'], 
                                                          f'{channel}.db'))
                        c = conn.cursor()
                        c.execute('''UPDATE messages SET media_path = ? WHERE message_id = ?''', 
                                (media_path, message.id))
                        conn.commit()
                        conn.close()
                
                last_message_id = message.id
                processed_messages += 1

                progress = (processed_messages / total_messages) * 100
                sys.stdout.write(f"\rScraping channel: {state['channels'][channel]['folder_name']} - Progress: {progress:.2f}%")
                sys.stdout.flush()

                state['channels'][channel]['last_id'] = last_message_id
                save_channels(state['channels'])
            except Exception as e:
                print(f"\nError processing message {message.id}: {str(e)}")
                continue
        print("\nScraping completed!")
    except Exception as e:
        print(f"\nError with channel {channel}: {str(e)}")

async def continuous_scraping():
    global continuous_scraping_active, scraping_log
    scraping_log = []
    start_time = datetime.now()
    messages_processed = 0
    channels_checked = 0
    
    try:
        while continuous_scraping_active:
            cycle_start = datetime.now()
            cycle_messages = 0
            print("\nChecking channels for new messages...")
            
            for channel in state['channels']:
                channel_start_messages = messages_processed
                print(f"\nChecking channel: {state['channels'][channel]['folder_name']}")
                await scrape_channel(channel, state['channels'][channel]['last_id'])
                
                # Calculate new messages for this channel
                new_messages = messages_processed - channel_start_messages
                if new_messages > 0:
                    scraping_log.append(f"{datetime.now().strftime('%H:%M:%S')} - {state['channels'][channel]['folder_name']}: {new_messages} new messages")
                    cycle_messages += new_messages
                
                channels_checked += 1
            
            if cycle_messages > 0:
                scraping_log.append(f"--- Cycle completed: {cycle_messages} new messages found ---\n")
            
            print("\nWaiting 60 seconds before next check...")
            print("(You can use other menu options while continuous scraping runs)")
            for i in range(60):
                if not continuous_scraping_active:
                    break
                await asyncio.sleep(1)
                sys.stdout.write(f"\rNext check in {60-i} seconds...")
                sys.stdout.flush()
                
    except asyncio.CancelledError:
        pass
    finally:
        total_duration = datetime.now() - start_time
        hours = total_duration.total_seconds() / 3600
        
        print("\n\n=== Continuous Scraping Summary ===")
        print(f"Duration: {hours:.1f} hours")
        print(f"Channels checked: {channels_checked}")
        
        if scraping_log:
            print("\nLast 10 activities:")
            for log in scraping_log[-10:]:
                print(log)
        else:
            print("\nNo new messages were found during the scraping period.")
            
        continuous_scraping_active = False

async def setup_google_credentials():
    creds = None
    if os.path.exists(GOOGLE_CREDS_FILE):
        with open(GOOGLE_CREDS_FILE, 'rb') as token:
            creds = pickle.load(token)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
                # Save refreshed credentials
                with open(GOOGLE_CREDS_FILE, 'wb') as token:
                    pickle.dump(creds, token)
            except Exception as e:
                print(f"\nError refreshing credentials: {str(e)}")
                os.remove(GOOGLE_CREDS_FILE)  # Remove invalid credentials
                creds = None
        
        if not creds:
            print("\nTo export to Google Sheets, you need to set up Google API credentials.")
            print("\nSetup instructions:")
            print("1. Go to https://console.cloud.google.com/")
            print("2. Create a new project")
            print("3. Enable BOTH of these APIs:")
            print("   - Google Sheets API")
            print("   - Google Drive API")
            print("4. Go to 'OAuth consent screen':")
            print("   - Select 'External'")
            print("   - Fill in required information")
            print("5. Go to 'Credentials':")
            print("   - Create OAuth 2.0 Client ID")
            print("   - Application type: Desktop application")
            print("   - Download the client configuration file")
            print("\nIMPORTANT: Make sure BOTH APIs are enabled before continuing!")
            
            while True:
                try:
                    client_secret_path = input("\nEnter the path to your downloaded client_secret.json file (or 'c' to cancel): ").strip()
                    
                    if client_secret_path.lower() == 'c':
                        print("Setup cancelled.")
                        return None
                    
                    # Remove quotes if user copied path with quotes
                    client_secret_path = client_secret_path.strip('"\'')
                    
                    # Expand user directory if path starts with ~
                    if client_secret_path.startswith('~'):
                        client_secret_path = os.path.expanduser(client_secret_path)
                    
                    # Convert to absolute path
                    client_secret_path = os.path.abspath(client_secret_path)
                    
                    if not os.path.exists(client_secret_path):
                        print("Error: File not found! Please check the path and try again.")
                        continue
                        
                    try:
                        flow = InstalledAppFlow.from_client_secrets_file(client_secret_path, SCOPES)
                        creds = flow.run_local_server(port=0)
                        
                        # Test the credentials with both APIs
                        print("\nTesting API access...")
                        sheets_service = build('sheets', 'v4', credentials=creds)
                        drive_service = build('drive', 'v3', credentials=creds)
                        
                        # Test Drive API
                        drive_service.files().list(pageSize=1).execute()
                        
                        # Test Sheets API by creating a temporary spreadsheet
                        test_spreadsheet = {
                            'properties': {'title': 'Test Spreadsheet'}
                        }
                        test_sheet = sheets_service.spreadsheets().create(body=test_spreadsheet).execute()
                        
                        # Clean up the test spreadsheet
                        drive_service.files().delete(fileId=test_sheet['spreadsheetId']).execute()
                        
                        # Save the working credentials
                        with open(GOOGLE_CREDS_FILE, 'wb') as token:
                            pickle.dump(creds, token)
                            
                        print("\nGoogle API credentials successfully set up!")
                        break
                        
                    except Exception as e:
                        error_msg = str(e)
                        if "accessNotConfigured" in error_msg:
                            if "drive.googleapis.com" in error_msg:
                                print("\nError: Google Drive API is not enabled!")
                                print("Please enable it in the Google Cloud Console and try again.")
                            elif "sheets.googleapis.com" in error_msg:
                                print("\nError: Google Sheets API is not enabled!")
                                print("Please enable it in the Google Cloud Console and try again.")
                        else:
                            print(f"\nError setting up credentials: {error_msg}")
                            print("\nPlease make sure:")
                            print("1. Both Google Drive and Google Sheets APIs are enabled")
                            print("2. The OAuth consent screen is configured")
                            print("3. You have created OAuth 2.0 credentials")
                            print("4. You have authorized the application")
                        
                        retry = input("Would you like to try again? (y/n): ").lower()
                        if retry != 'y':
                            return None
                        
                except KeyboardInterrupt:
                    print("\nOperation cancelled.")
                    return None
                except Exception as e:
                    print(f"\nError: {str(e)}")
                    retry = input("Would you like to try again? (y/n): ").lower()
                    if retry != 'y':
                        return None
                    
    return creds

# Add this function to calculate file size
def get_file_size(file_path):
    """Get file size in bytes"""
    return os.path.getsize(file_path)

# Replace the get_upload_speed function with this version
def calculate_upload_speed(file_size, elapsed_time):
    """Calculate upload speed in MB/s"""
    if elapsed_time > 0:
        return (file_size / (1024 * 1024)) / elapsed_time  # Convert bytes to MB and calculate speed
    return 0

# Add this helper function to create or get folder
async def create_or_get_folder(drive_service, folder_name, parent_id=None):
    """Create a folder in Google Drive or get it if it exists"""
    try:
        # Check if folder already exists
        query = f"name='{folder_name}' and mimeType='application/vnd.google-apps.folder'"
        if parent_id:
            query += f" and '{parent_id}' in parents"
        
        results = drive_service.files().list(
            q=query,
            spaces='drive',
            fields='files(id, name)',
            supportsAllDrives=True
        ).execute()
        
        if results['files']:
            return results['files'][0]['id']
        
        # Create folder if it doesn't exist
        folder_metadata = {
            'name': folder_name,
            'mimeType': 'application/vnd.google-apps.folder'
        }
        if parent_id:
            folder_metadata['parents'] = [parent_id]
        
        folder = drive_service.files().create(
            body=folder_metadata,
            fields='id',
            supportsAllDrives=True
        ).execute()
        
        # Set folder permissions
        try:
            drive_service.permissions().create(
                fileId=folder.get('id'),
                body={
                    'type': 'anyone',
                    'role': 'writer',  # Change to writer instead of reader
                    'allowFileDiscovery': True
                },
                fields='id',
                supportsAllDrives=True
            ).execute()
        except Exception as e:
            print(f"Warning: Could not set permissions for folder {folder_name}: {str(e)}")
        
        return folder.get('id')
    except Exception as e:
        print(f"Error creating/getting folder {folder_name}: {str(e)}")
        raise

# Add this function to store and retrieve folder IDs
def save_folder_ids(channel, folder_ids):
    """Save folder IDs to channel state"""
    if 'folder_ids' not in state['channels'][channel]:
        state['channels'][channel]['folder_ids'] = {}
    state['channels'][channel]['folder_ids'].update(folder_ids)
    save_channels(state['channels'])

def get_folder_ids(channel):
    """Get saved folder IDs for channel"""
    return state['channels'][channel].get('folder_ids', {})

# Add this helper function to get folder link
def get_folder_link(folder_id):
    """Get shareable link for a Google Drive folder"""
    return f"https://drive.google.com/drive/folders/{folder_id}"

# Add this helper function to get formatted channel folder name
def get_channel_folder_name(channel_name, last_scrape=None):
    """Get channel folder name with last scrape date"""
    if last_scrape:
        return f"{channel_name} (Last Scrape: {last_scrape})"
    return channel_name

# Add this helper function to chunk data
def chunk_data(data, chunk_size=100):
    """Split data into smaller chunks"""
    return [data[i:i + chunk_size] for i in range(0, len(data), chunk_size)]

# Add this async function for uploading to sheets
async def upload_to_sheets(sheets_service, spreadsheet_id, values, columns):
    try:
        total_rows = len(values)
        chunk_size = 1000  # Google Sheets API limit per request
        
        # First, resize the sheet to accommodate all data
        print("\nResizing spreadsheet...")
        requests = [
            {
                'updateSheetProperties': {
                    'properties': {
                        'sheetId': 0,  # First sheet
                        'gridProperties': {
                            'rowCount': total_rows + 100,  # Add some buffer
                            'columnCount': len(columns) + 1  # Add buffer for media column
                        }
                    },
                    'fields': 'gridProperties(rowCount,columnCount)'
                }
            }
        ]
        
        sheets_service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={'requests': requests}
        ).execute()
        
        # Upload data in chunks
        print(f"\nUploading {total_rows} rows in {math.ceil(total_rows/chunk_size)} chunks...")
        with tqdm(total=math.ceil(total_rows/chunk_size), desc="Uploading to spreadsheet") as pbar:
            for i in range(0, total_rows, chunk_size):
                chunk = values[i:i + chunk_size]
                range_name = f'A{i+1}' if i == 0 else f'A{i+1}:Z{i+len(chunk)}'
                
                max_retries = 3
                for attempt in range(max_retries):
                    try:
                        sheets_service.spreadsheets().values().update(
                            spreadsheetId=spreadsheet_id,
                            range=range_name,
                            valueInputOption='USER_ENTERED',
                            body={'values': chunk}
                        ).execute()
                        break
                    except Exception as e:
                        if attempt == max_retries - 1:
                            print(f"\nFailed to upload chunk {i//chunk_size + 1}/{math.ceil(total_rows/chunk_size)} after {max_retries} attempts")
                            raise e
                        print(f"\nRetrying chunk {i//chunk_size + 1} ({attempt + 2}/{max_retries})...")
                        await asyncio.sleep(2)  # Wait before retrying
                
                pbar.update(1)
        
        return True
        
    except Exception as e:
        print(f"Error during spreadsheet upload: {e}")
        return False

# Add this helper function to check for existing media
async def get_existing_media(drive_service, folder_id):
    """Get a map of existing media files in the folder"""
    existing_media = {}
    page_token = None
    
    print("\nChecking existing media files...")
    while True:
        try:
            # Query files in the media folder
            query = f"'{folder_id}' in parents and trashed = false"
            response = drive_service.files().list(
                q=query,
                spaces='drive',
                fields='nextPageToken, files(id, name)',
                pageToken=page_token
            ).execute()
            
            # Add files to our map
            for file in response.get('files', []):
                existing_media[file['name']] = file['id']
                
            # Check if there are more pages
            page_token = response.get('nextPageToken')
            if not page_token:
                break
                
        except Exception as e:
            print(f"Error checking existing media: {str(e)}")
            break
    
    print(f"Found {len(existing_media)} existing media files")
    return existing_media

# Add this helper function to handle multiple media files
def get_media_links(media_files):
    """Convert list of media links to a single string with links stacked vertically"""
    if not media_files:
        return ""
    return "\n".join(media_files)

async def export_to_google_sheets(channel, creds):
    try:
        channel_dir = os.path.join(os.getcwd(), state['channels'][channel]['folder_name'])
        db_file = os.path.join(channel_dir, f'{channel}.db')
        
        if not os.path.exists(db_file):
            raise FileNotFoundError(f"Database file not found for channel {channel}")
        
        print("\nInitializing Google Services...")
        sheets_service = build('sheets', 'v4', credentials=creds)
        drive_service = build('drive', 'v3', credentials=creds)
        
        # Get existing folder IDs and create folder structure
        folder_ids = get_folder_ids(channel)
        
        # Verify and create folder structure
        print("\nChecking folder structure...")
        with tqdm(total=4, desc="Setting up folders", unit="folder") as pbar:
            # Main Telegram Scraper folder
            if 'main_folder' not in folder_ids:
                main_folder_id = await create_or_get_folder(drive_service, "Telegram Scraper")
                folder_ids['main_folder'] = main_folder_id
            else:
                main_folder_id = folder_ids['main_folder']
            pbar.update(1)
            
            # Channel folder
            if 'channel_folder' not in folder_ids:
                channel_folder_name = get_channel_folder_name(
                    state['channels'][channel]['folder_name'],
                    datetime.now().strftime('%Y-%m-%d %H:%M')
                )
                channel_folder_id = await create_or_get_folder(drive_service, channel_folder_name, main_folder_id)
                folder_ids['channel_folder'] = channel_folder_id
            else:
                channel_folder_id = folder_ids['channel_folder']
                # Update existing channel folder name with new scrape date
                new_folder_name = get_channel_folder_name(
                    state['channels'][channel]['folder_name'],
                    datetime.now().strftime('%Y-%m-%d %H:%M')
                )
                try:
                    drive_service.files().update(
                        fileId=channel_folder_id,
                        body={'name': new_folder_name}
                    ).execute()
                except Exception:
                    # If update fails, create new folder
                    channel_folder_id = await create_or_get_folder(drive_service, new_folder_name, main_folder_id)
                    folder_ids['channel_folder'] = channel_folder_id
            pbar.update(1)
            
            # Google Sheets folder
            if 'sheets_folder' not in folder_ids:
                sheets_folder_id = await create_or_get_folder(drive_service, "Google Sheets", channel_folder_id)
                folder_ids['sheets_folder'] = sheets_folder_id
            else:
                sheets_folder_id = folder_ids['sheets_folder']
            pbar.update(1)
            
            # Media folder
            if 'media_folder' not in folder_ids:
                media_folder_id = await create_or_get_folder(drive_service, "Media", channel_folder_id)
                folder_ids['media_folder'] = media_folder_id
            else:
                media_folder_id = folder_ids['media_folder']
            pbar.update(1)
            
            # Save folder IDs for future use
            save_folder_ids(channel, folder_ids)
            pbar.update(4)
        
        # Create new spreadsheet
        print("\nCreating spreadsheet...")
        spreadsheet = {
            'properties': {
                'title': f"{state['channels'][channel]['folder_name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            }
        }
        spreadsheet = sheets_service.spreadsheets().create(body=spreadsheet).execute()
        spreadsheet_id = spreadsheet['spreadsheetId']
        
        # Move spreadsheet to sheets folder
        file = drive_service.files().get(fileId=spreadsheet_id, fields='parents').execute()
        previous_parents = ",".join(file.get('parents'))
        drive_service.files().update(
            fileId=spreadsheet_id,
            addParents=sheets_folder_id,
            removeParents=previous_parents,
            fields='id, parents'
        ).execute()
        
        # Get data from database
        conn = sqlite3.connect(db_file)
        c = conn.cursor()
        c.execute('SELECT * FROM messages')
        rows = c.fetchall()
        columns = [description[0] for description in c.description]
        columns.append('media_preview')
        
        # Get existing media files
        existing_media = await get_existing_media(drive_service, media_folder_id)
        
        # First pass: identify main message rows and collect all media
        message_media_map = {}
        message_rows = {}
        message_to_media_id = {}  # New map to track which media belongs to which message

        # First, get the parent message for each media
        for row in rows:
            message_id = row[columns.index('message_id')]
            reply_to = row[columns.index('reply_to')] if 'reply_to' in columns else None
            media_path = row[columns.index('media_path')]
            message_text = row[columns.index('message')]
            
            # Safely handle None values for message text
            message_text = message_text.strip() if message_text else ""
            
            # If this is a media-only row, find its parent message
            if not message_text and media_path:
                if reply_to:
                    message_to_media_id[message_id] = reply_to
                else:
                    # If no reply_to, look for nearest message before this one
                    c.execute('''
                        SELECT message_id FROM messages 
                        WHERE message_id < ? AND message IS NOT NULL AND message != '' 
                        ORDER BY message_id DESC LIMIT 1
                    ''', (message_id,))
                    result = c.fetchone()
                    if result:
                        message_to_media_id[message_id] = result[0]

        # Now collect messages and media
        for row in rows:
            message_id = row[columns.index('message_id')]
            media_path = row[columns.index('media_path')]
            message_text = row[columns.index('message')]
            
            # Safely handle None values
            message_text = message_text.strip() if message_text else ""
            
            # Store rows with actual messages
            if message_text:
                message_rows[message_id] = list(row)
            
            # Determine which message this media belongs to
            target_message_id = message_to_media_id.get(message_id, message_id)
            
            # Collect all media paths
            if media_path and os.path.exists(os.path.join(channel_dir, media_path)):
                if target_message_id not in message_media_map:
                    message_media_map[target_message_id] = []
                if media_path not in message_media_map[target_message_id]:  # Avoid duplicates
                    message_media_map[target_message_id].append(media_path)

        # Sort messages by media count
        sorted_message_ids = sorted(
            message_rows.keys(),
            key=lambda msg_id: len(message_media_map.get(msg_id, [])),
            reverse=True  # Reverse to put messages with most media first
        )

        # Now process only the rows with messages, in sorted order
        values = [columns]  # Headers
        with tqdm(total=len(message_rows), desc="Processing rows") as pbar:
            for message_id in sorted_message_ids:
                row_data = message_rows[message_id]
                media_files = []
                media_type = row_data[columns.index('media_type')] if 'media_type' in columns else None
                
                # Process all media files for this message
                if message_id in message_media_map:
                    for media_path in message_media_map[message_id]:
                        file_name = os.path.basename(media_path)
                        
                        if file_name in existing_media:
                            file_id = existing_media[file_name]
                            print(f"\nUsing existing media file: {file_name}")
                        else:
                            try:
                                full_path = os.path.join(channel_dir, media_path)
                                file_metadata = {
                                    'name': file_name,
                                    'parents': [media_folder_id]
                                }
                                media = MediaFileUpload(full_path, resumable=True)
                                file = drive_service.files().create(
                                    body=file_metadata,
                                    media_body=media,
                                    fields='id'
                                ).execute()
                                file_id = file.get('id')
                                existing_media[file_name] = file_id
                                print(f"\nUploaded new media file: {file_name}")
                            except Exception as e:
                                print(f"\nError uploading media: {str(e)}")
                                continue
                        
                        drive_link = f"https://drive.google.com/file/d/{file_id}/view"
                        media_files.append(drive_link)
                
                elif media_type == 'MessageMediaPoll':
                    message_text = row_data[columns.index('message')]
                    media_files.append(f"Poll: {message_text}")
                
                # Add all media links to the row
                row_data.append(get_media_links(media_files))
                values.append(row_data)
                pbar.update(1)

        conn.close()
        
        # Upload data to spreadsheet
        print("\nUploading data to spreadsheet...")
        await upload_to_sheets(sheets_service, spreadsheet_id, values, columns)
        
        # Format spreadsheet
        print("\nFormatting spreadsheet...")
        requests = [
            {
                'updateDimensionProperties': {
                    'range': {'sheetId': 0, 'dimension': 'COLUMNS'},
                    'properties': {'pixelSize': 200},
                    'fields': 'pixelSize'
                }
            },
            {
                'repeatCell': {
                    'range': {'startRowIndex': 0, 'endRowIndex': 1},
                    'cell': {
                        'userEnteredFormat': {
                            'backgroundColor': {'red': 0.2, 'green': 0.4, 'blue': 0.7},
                            'textFormat': {'bold': True, 'foregroundColor': {'red': 1.0, 'green': 1.0, 'blue': 1.0}}
                        }
                    },
                    'fields': 'userEnteredFormat(backgroundColor,textFormat)'
                }
            }
        ]
        
        sheets_service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={'requests': requests}
        ).execute()
        
        return True
        
    except Exception as e:
        print(f"Error exporting to Google Sheets: {str(e)}")
        return False

# Modify the export_data function
async def export_data():
    if not state['channels']:
        print("\nNo channels to export. Please add and scrape channels first.")
        return
        
    print("\nSelect export format:")
    print("[1] Local Excel/JSON")
    print("[2] Google Sheets")
    print("[3] Both")
    print("[C] Cancel")
    
    format_choice = input("\nEnter your choice: ").strip().lower()
    
    if format_choice == 'c':
        print("Operation cancelled.")
        return
        
    if format_choice not in ['1', '2', '3']:
        print("Invalid choice.")
        return
    
    # Setup Google credentials if needed
    creds = None
    if format_choice in ['2', '3']:
        creds = await setup_google_credentials()
        if not creds:
            print("\nFailed to set up Google credentials.")
            if format_choice == '2':
                return
            format_choice = '1'  # Fall back to local export if credentials fail
        
    # Get channel selection
    print("\nSelect channels to export:")
    channels_list = list(state['channels'].items())
    
    for idx, (channel_id, channel_info) in enumerate(channels_list, 1):
        print(f"[{idx}] {channel_info['folder_name']}")
    
    while True:
        selection = input("\nEnter channel numbers to export (comma-separated, 'all' for all channels, or 'c' to cancel): ").strip().lower()
        
        if selection == 'c':
            print("Operation cancelled.")
            return
            
        if selection == 'all':
            selected_channels = [channel_id for channel_id, _ in channels_list]
        else:
            try:
                selected_indices = [int(idx.strip()) for idx in selection.split(',')]
                if all(1 <= idx <= len(channels_list) for idx in selected_indices):
                    selected_channels = [channels_list[idx-1][0] for idx in selected_indices]
                else:
                    print("Invalid selection. Please enter numbers within the range shown.")
                    continue
            except ValueError:
                print("Invalid input. Please enter numbers separated by commas (e.g., 1,3,4)")
                continue
        
        print("\nExporting data...")
        exported_count = 0
        for channel in selected_channels:
            try:
                if format_choice in ['1', '3']:
                    export_to_excel(channel)
                    export_to_json(channel)
                    
                if format_choice in ['2', '3'] and creds:
                    await export_to_google_sheets(channel, creds)
                    
                exported_count += 1
                print(f"Successfully exported data for {state['channels'][channel]['folder_name']}")
            except Exception as e:
                print(f"Error exporting data for {state['channels'][channel]['folder_name']}: {str(e)}")
        
        print(f"\nExport completed! Exported data for {exported_count} channel(s).")
        break

def export_to_csv(channel):
    channel_dir = os.path.join(os.getcwd(), channel)
    db_file = os.path.join(channel_dir, f'{channel}.db')
    csv_file = os.path.join(channel_dir, f'{channel}.csv')
    
    if not os.path.exists(db_file):
        raise FileNotFoundError(f"Database file not found for channel {channel}")
        
    conn = sqlite3.connect(db_file)
    c = conn.cursor()
    c.execute('SELECT * FROM messages')
    rows = c.fetchall()
    
    if not rows:
        print(f"No messages found in channel {channel}")
        conn.close()
        return
        
    with open(csv_file, 'w', newline='', encoding='utf-8-sig') as f:  # Changed to utf-8-sig for better Excel compatibility
        writer = csv.writer(f)
        writer.writerow([description[0] for description in c.description])
        writer.writerows(rows)
    conn.close()

def export_to_json(channel):
    # Create base channel directory
    channel_dir = os.path.join(os.getcwd(), state['channels'][channel]['folder_name'])
    # Create exports directory
    exports_dir = os.path.join(channel_dir, 'exports')
    os.makedirs(exports_dir, exist_ok=True)
    
    db_file = os.path.join(channel_dir, f'{channel}.db')
    # Put json file in exports directory with timestamp
    json_file = os.path.join(exports_dir, f'{channel}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json')
    
    if not os.path.exists(db_file):
        raise FileNotFoundError(f"Database file not found for channel {channel}")
        
    conn = sqlite3.connect(db_file)
    c = conn.cursor()
    c.execute('SELECT * FROM messages')
    rows = c.fetchall()
    
    if not rows:
        print(f"No messages found in channel {channel}")
        conn.close()
        return
        
    data = []
    columns = [description[0] for description in c.description]
    for row in rows:
        item = dict(zip(columns, row))
        # Parse JSON fields
        for field in ['entities', 'reactions']:
            if item[field]:
                try:
                    item[field] = json.loads(item[field])
                except:
                    pass
        data.append(item)
        
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)
    conn.close()

def create_thumbnail(image_path, max_size=(100, 100)):
    """Create a thumbnail while maintaining aspect ratio"""
    try:
        with PILImage.open(image_path) as img:
            # Calculate thumbnail size maintaining aspect ratio
            thumb = img.copy()
            thumb.thumbnail(max_size, PILImage.Resampling.LANCZOS)
            
            # Save thumbnail to BytesIO
            bio = BytesIO()
            thumb.save(bio, format='PNG')
            bio.seek(0)
            return bio
    except Exception as e:
        print(f"Error creating thumbnail: {str(e)}")
        return None

def export_to_excel(channel):
    # Create base channel directory
    channel_dir = os.path.join(os.getcwd(), state['channels'][channel]['folder_name'])
    # Create exports directory
    exports_dir = os.path.join(channel_dir, 'exports')
    os.makedirs(exports_dir, exist_ok=True)
    
    db_file = os.path.join(channel_dir, f'{channel}.db')
    # Put excel file in exports directory
    excel_file = os.path.join(exports_dir, f'{channel}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    
    if not os.path.exists(db_file):
        raise FileNotFoundError(f"Database file not found for channel {channel}")
        
    conn = sqlite3.connect(db_file)
    c = conn.cursor()
    c.execute('SELECT * FROM messages')
    rows = c.fetchall()
    columns = [description[0] for description in c.description]
    
    if not rows:
        print(f"No messages found in channel {channel}")
        conn.close()
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Messages"

    # Add a new column for image thumbnails
    columns.append('thumbnail')
    
    # Style for headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Write headers
    for col, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col)].width = max(len(header) + 2, 15)

    # Create AppleScript for image interaction
    applescript = f'''
    tell application "Microsoft Excel"
        tell active sheet
            set imageShape to (get shape "{{image_name}}")
            if height of imageShape is less than or equal to 75 then
                set height of imageShape to original height of imageShape
                set width of imageShape to original width of imageShape
            else
                set aspectRatio to (width of imageShape / height of imageShape)
                if aspectRatio > 1 then
                    set height of imageShape to 75 / aspectRatio
                    set width of imageShape to 75
                else
                    set height of imageShape to 75
                    set width of imageShape to aspectRatio * 75
                end if
            end if
        end tell
    end tell
    '''

    # Write data and add images
    media_col_idx = columns.index('media_path') + 1
    thumbnail_col_idx = len(columns)
    row_height = 75

    print("\nProcessing messages and adding images...")
    total_rows = len(rows)
    
    for row_idx, row in enumerate(rows, 2):
        if row_idx % 10 == 0:
            progress = (row_idx / total_rows) * 100
            sys.stdout.write(f"\rProcessing row {row_idx}/{total_rows} ({progress:.1f}%)...")
            sys.stdout.flush()

        for col_idx, value in enumerate(row, 1):
            if columns[col_idx-1] in ['entities', 'reactions'] and value:
                try:
                    parsed_value = json.loads(value)
                    value = json.dumps(parsed_value, indent=2, ensure_ascii=False)
                except:
                    pass
            
            # Handle media paths
            if col_idx == media_col_idx and value:
                media_path = os.path.join(channel_dir, value)
                if os.path.exists(media_path) and any(media_path.lower().endswith(ext) for ext in ['.jpg', '.jpeg', '.png']):
                    try:
                        # Create full-size image and thumbnail
                        img = Image(media_path)
                        cell = ws.cell(row=row_idx, column=thumbnail_col_idx)
                        
                        # Calculate thumbnail size
                        original_width = img.width
                        original_height = img.height
                        aspect_ratio = original_width / original_height
                        
                        if aspect_ratio > 1:
                            new_height = 75 / aspect_ratio
                            new_width = 75
                        else:
                            new_height = 75
                            new_width = 75 * aspect_ratio
                        
                        img.width = new_width
                        img.height = new_height
                        img.anchor = cell.coordinate
                        
                        # Add image with unique name for AppleScript reference
                        image_name = f"img_{row_idx}"
                        img.name = image_name
                        ws.add_image(img)
                        
                        # Create AppleScript file for this image
                        script_path = os.path.join(channel_dir, f"{image_name}.scpt")
                        with open(script_path, 'w') as f:
                            f.write(applescript.format(image_name=image_name))
                        
                        # Make script executable
                        os.chmod(script_path, 0o755)
                        
                    except Exception as e:
                        print(f"\nError adding image for row {row_idx}: {str(e)}")
            
            ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else '')

    # Auto-adjust column widths
    for col in ws.columns:
        if get_column_letter(col[0].column) != get_column_letter(thumbnail_col_idx):
            max_length = 0
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 50)

    print("\nSaving Excel file...")
    wb.save(excel_file)
    conn.close()
    
    # Update README path to exports directory
    readme_path = os.path.join(exports_dir, "README.txt")
    with open(readme_path, 'w') as f:
        f.write("""
Image Interaction Instructions:
1. Double-click on an image to view it in full size
2. To resize images:
   - Select the image
   - Right-click and choose 'Format Picture'
   - Adjust size as needed
   
Note: Images are saved in the media folder at full resolution.
""")
    
    print(f"Excel file exported: {excel_file}")
    print(f"Please read {readme_path} for image interaction instructions")

async def view_channels():
    if not state['channels']:
        print("No channels to view.")
        return
    
    print("\nCurrent channels:")
    for channel, last_id in state['channels'].items():
        print(f"Channel ID: {channel}, Last Message ID: {last_id}")

async def list_Channels():
    try:
        print("\nList of channels joined by account: ")
        async for dialog in client.iter_dialogs():
            if (dialog.id != 777000):
                print(f"* {dialog.title} (id: {dialog.id})")
    except Exception as e:
        print(f"Error processing: {e}")

async def select_channels_to_scrape():
    if not state['channels']:
        print("\nNo channels saved. Please add channels first.")
        return []

    print("\nSaved channels:")
    channels_list = list(state['channels'].items())
    
    for idx, (channel_id, channel_info) in enumerate(channels_list, 1):
        print(f"[{idx}] {channel_info['folder_name']}")
    
    while True:
        selection = input("\nEnter channel numbers to scrape (comma-separated, or 'all' for all channels): ").strip().lower()
        
        if selection == 'all':
            return [channel_id for channel_id, _ in channels_list]
        
        try:
            # Split by comma and remove whitespace
            selected_indices = [int(idx.strip()) for idx in selection.split(',')]
            
            # Validate indices
            if all(1 <= idx <= len(channels_list) for idx in selected_indices):
                # Convert indices to channel IDs
                selected_channels = [channels_list[idx-1][0] for idx in selected_indices]
                return selected_channels
            else:
                print("Invalid selection. Please enter numbers within the range shown.")
        except ValueError:
            print("Invalid input. Please enter numbers separated by commas (e.g., 1,3,4)")

async def add_from_channel_list():
    try:
        print("\nList of channels you can add:")
        channels = []
        idx = 1
        
        async for dialog in client.iter_dialogs():
            if dialog.id != 777000:  # Skip Telegram service notifications
                channels.append((dialog.id, dialog.title))
                print(f"[{idx}] {dialog.title} (id: {dialog.id})")
                idx += 1
        
        if not channels:
            print("No channels found!")
            return
            
        while True:
            selection = input("\nEnter channel numbers to add (comma-separated, 'all' for all channels, or 'c' to cancel): ").strip().lower()
            
            if selection == 'c':
                print("Operation cancelled.")
                return
                
            if selection == 'all':
                selected_channels = channels
            else:
                try:
                    # Split by comma and remove whitespace
                    selected_indices = [int(idx.strip()) for idx in selection.split(',')]
                    
                    # Validate indices
                    if all(1 <= idx <= len(channels) for idx in selected_indices):
                        selected_channels = [channels[idx-1] for idx in selected_indices]
                    else:
                        print("Invalid selection. Please enter numbers within the range shown.")
                        continue
                except ValueError:
                    print("Invalid input. Please enter numbers separated by commas (e.g., 1,3,4)")
                    continue
            
            # Add selected channels
            added_count = 0
            for channel_id, channel_title in selected_channels:
                if str(channel_id) not in state['channels']:
                    folder_name = f"{channel_title}_{channel_id}"
                    state['channels'][str(channel_id)] = {
                        'last_id': 0,
                        'folder_name': folder_name
                    }
                    added_count += 1
                    print(f"Added channel: {channel_title}")
                else:
                    print(f"Channel {channel_title} is already in saved channels")
            
            save_channels(state['channels'])
            print(f"\nAdded {added_count} new channel(s) to saved channels.")
            break
            
    except Exception as e:
        print(f"Error processing: {str(e)}")

async def remove_channels():
    if not state['channels']:
        print("\nNo channels to remove.")
        return
        
    print("\nSelect channels to remove:")
    channels_list = list(state['channels'].items())
    
    for idx, (channel_id, channel_info) in enumerate(channels_list, 1):
        print(f"[{idx}] {channel_info['folder_name']}")
    
    while True:
        selection = input("\nEnter channel numbers to remove (comma-separated, 'all' for all channels, or 'c' to cancel): ").strip().lower()
        
        if selection == 'c':
            print("Operation cancelled.")
            return
            
        if selection == 'all':
            confirm = input("Are you sure you want to remove ALL channels? (y/n): ").strip().lower()
            if confirm == 'y':
                state['channels'].clear()
                save_channels(state['channels'])
                print("All channels removed.")
            else:
                print("Operation cancelled.")
            return
            
        try:
            # Split by comma and remove whitespace
            selected_indices = [int(idx.strip()) for idx in selection.split(',')]
            
            # Validate indices
            if all(1 <= idx <= len(channels_list) for idx in selected_indices):
                # Remove selected channels
                removed_count = 0
                for idx in sorted(selected_indices, reverse=True):  # Remove in reverse order to maintain indices
                    channel_id, channel_info = channels_list[idx-1]
                    del state['channels'][channel_id]
                    print(f"Removed channel: {channel_info['folder_name']}")
                    removed_count += 1
                
                save_channels(state['channels'])
                print(f"\nRemoved {removed_count} channel(s).")
                break
            else:
                print("Invalid selection. Please enter numbers within the range shown.")
        except ValueError:
            print("Invalid input. Please enter numbers separated by commas (e.g., 1,3,4)")

async def validate_channel_id(channel_id):
    try:
        if channel_id.startswith('-'):
            entity = await client.get_entity(PeerChannel(int(channel_id)))
        else:
            entity = await client.get_entity(channel_id)
        return True, entity.title
    except ValueError:
        return False, "Invalid channel ID format"
    except Exception as e:
        return False, str(e)

async def scrape_new_messages(channel, last_id):
    """Scrape only new messages since last scrape"""
    try:
        if channel not in state['channels']:
            print(f"\nChannel {channel} not found in saved channels.")
            return

        if channel.startswith('-'):
            entity = await client.get_entity(PeerChannel(int(channel)))
        else:
            entity = await client.get_entity(channel)

        print(f"\nChecking for new messages in {state['channels'][channel]['folder_name']}...")
        total_new = 0
        async for _ in client.iter_messages(entity, min_id=last_id):
            total_new += 1

        if total_new == 0:
            print("No new messages found.")
            return

        print(f"Found {total_new} new messages to process")
        processed = 0

        async for message in client.iter_messages(entity, min_id=last_id):
            try:
                sender = await message.get_sender()
                save_message_to_db(channel, message, sender)

                if state['scrape_media'] and message.media:
                    media_path = await download_media(channel, message)
                    if media_path:
                        conn = sqlite3.connect(os.path.join(os.getcwd(), 
                                                          state['channels'][channel]['folder_name'], 
                                                          f'{channel}.db'))
                        c = conn.cursor()
                        c.execute('''UPDATE messages SET media_path = ? WHERE message_id = ?''', 
                                (media_path, message.id))
                        conn.commit()
                        conn.close()

                processed += 1
                progress = (processed / total_new) * 100
                sys.stdout.write(f"\rProgress: {progress:.2f}% ({processed}/{total_new} messages)")
                sys.stdout.flush()

                # Update last message ID
                state['channels'][channel]['last_id'] = max(message.id, state['channels'][channel]['last_id'])
                save_channels(state['channels'])

            except Exception as e:
                print(f"\nError processing message {message.id}: {str(e)}")
                continue

        print(f"\nCompleted! Scraped {processed} new messages.")

    except Exception as e:
        print(f"\nError with channel {channel}: {str(e)}")

async def delete_channel_data():
    """Delete all saved data for selected channels including directories and logs"""
    if not state['channels']:
        print("\nNo channels to delete data from.")
        return
        
    print("\nSelect channels to delete data from:")
    channels_list = list(state['channels'].items())
    
    for idx, (channel_id, channel_info) in enumerate(channels_list, 1):
        print(f"[{idx}] {channel_info['folder_name']}")
    
    while True:
        selection = input("\nEnter channel numbers to delete data from (comma-separated, 'all' for all channels, or 'c' to cancel): ").strip().lower()
        
        if selection == 'c':
            print("Operation cancelled.")
            return
            
        if selection == 'all':
            confirm = input("Are you sure you want to delete ALL channel data? This cannot be undone! (y/n): ").strip().lower()
            if confirm == 'y':
                selected_channels = channels_list
            else:
                print("Operation cancelled.")
                return
        else:
            try:
                selected_indices = [int(idx.strip()) for idx in selection.split(',')]
                
                if all(1 <= idx <= len(channels_list) for idx in selected_indices):
                    selected_channels = [channels_list[idx-1] for idx in selected_indices]
                else:
                    print("Invalid selection. Please enter numbers within the range shown.")
                    continue
            except ValueError:
                print("Invalid input. Please enter numbers separated by commas (e.g., 1,3,4)")
                continue
        
        # Confirm deletion
        channel_names = [channel_info['folder_name'] for _, channel_info in selected_channels]
        print("\nYou are about to delete all data for these channels:")
        for name in channel_names:
            print(f"- {name}")
        
        confirm = input("\nThis action cannot be undone! Are you sure? (y/n): ").strip().lower()
        if confirm != 'y':
            print("Operation cancelled.")
            return
        
        # Delete data for selected channels
        deleted_count = 0
        for channel_id, channel_info in selected_channels:
            try:
                # Get channel directory path
                channel_dir = os.path.join(os.getcwd(), channel_info['folder_name'])
                
                # Delete the channel directory and all its contents
                if os.path.exists(channel_dir):
                    import shutil
                    shutil.rmtree(channel_dir)
                    print(f"Deleted directory: {channel_dir}")
                
                # Reset the channel's last_id in state
                state['channels'][channel_id]['last_id'] = 0
                deleted_count += 1
                
            except Exception as e:
                print(f"Error deleting data for {channel_info['folder_name']}: {str(e)}")
        
        # Save updated state
        save_channels(state['channels'])
        print(f"\nSuccessfully deleted data for {deleted_count} channel(s).")
        break

# Modify the scrape_date_range function
async def scrape_date_range(channel, start_date, end_date):
    """Scrape messages within a specific date range"""
    try:
        if channel.startswith('-'):
            entity = await client.get_entity(PeerChannel(int(channel)))
        else:
            entity = await client.get_entity(channel)

        print(f"\nCounting messages in date range for {state['channels'][channel]['folder_name']}...")
        
        # Convert dates to datetime objects with UTC timezone
        start_datetime = datetime.combine(start_date, datetime.min.time()).replace(tzinfo=timezone.utc)
        end_datetime = datetime.combine(end_date, datetime.max.time()).replace(tzinfo=timezone.utc)
        
        # First pass: get all messages and filter by date
        valid_message_ids = []
        
        print("\nSearching for messages in date range...")
        async for message in client.iter_messages(entity, limit=None):
            try:
                if start_datetime <= message.date <= end_datetime:
                    valid_message_ids.append(message.id)
            except Exception as e:
                print(f"\nError checking message date: {str(e)}")
                continue

        total_messages = len(valid_message_ids)
        if total_messages == 0:
            print("No messages found in the specified date range.")
            return

        print(f"\nFound {total_messages} messages in date range")
        processed_messages = 0

        # Second pass: fetch and process only the valid messages
        print("\nScraping messages...")
        with tqdm(total=total_messages, desc="Processing messages", unit="msg") as pbar:
            for message_id in valid_message_ids:
                try:
                    # Get specific message by ID
                    message = await client.get_messages(entity, ids=message_id)
                    if not message:
                        continue

                    sender = await message.get_sender()
                    save_message_to_db(channel, message, sender)

                    if state['scrape_media'] and message.media:
                        media_path = await download_media(channel, message)
                        if media_path:
                            conn = sqlite3.connect(os.path.join(os.getcwd(), 
                                                              state['channels'][channel]['folder_name'], 
                                                              f'{channel}.db'))
                            c = conn.cursor()
                            c.execute('''UPDATE messages SET media_path = ? WHERE message_id = ?''', 
                                    (media_path, message.id))
                            conn.commit()
                            conn.close()

                    processed_messages += 1
                    pbar.update(1)

                except Exception as e:
                    print(f"\nError processing message {message_id}: {str(e)}")
                    continue

        print(f"\nCompleted! Scraped {processed_messages} messages from {start_date} to {end_date}")

    except Exception as e:
        print(f"\nError with channel {channel}: {str(e)}")

# Add this function to handle date range input
def get_date_range():
    """Get start and end dates from user input"""
    while True:
        try:
            print("\nEnter date range (format: YYYY-MM-DD)")
            start_date_str = input("Start date: ").strip()
            end_date_str = input("End date: ").strip()
            
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
            
            if start_date > end_date:
                print("Start date must be before or equal to end date!")
                continue
                
            return start_date, end_date
            
        except ValueError:
            print("Invalid date format! Please use YYYY-MM-DD")
            continue

async def manage_channels():
    while True:
        print("\nMenu:")
        print("[A] Add new channel by ID")
        print("[L] List and add channels")
        print("[R] Remove channels")
        print("[S] Scrape selected channels (full rescrape)")
        print("[N] Scrape new messages only")
        print("[M] Toggle media scraping (currently {})".format(
            "enabled" if state['scrape_media'] else "disabled"))
        print("[C] Start continuous scraping")
        print("[X] Stop continuous scraping")
        print("[E] Export data")
        print("[V] View saved channels")
        print("[D] Delete channel data")
        print("[T] Scrape by date range")
        print("[O] Logout")
        print("[Q] Quit")

        choice = input("Enter your choice: ").lower()
        match (choice):
            case 'a':
                channel = input("Enter channel ID: ")
                is_valid, message = await validate_channel_id(channel)
                if is_valid:
                    folder_name = await get_channel_name(channel)
                    state['channels'][channel] = {
                        'last_id': 0,
                        'folder_name': folder_name
                    }
                    save_channels(state['channels'])
                    print(f"Added channel: {folder_name}")
                else:
                    print(f"Error: Could not add channel. {message}")
            case 'r':
                await remove_channels()
            case 's':
                selected_channels = await select_channels_to_scrape()
                if selected_channels:
                    print(f"\nScraping {len(selected_channels)} channel(s)...")
                    for channel in selected_channels:
                        # Pass force_rescrape=True to always scrape from beginning
                        await scrape_channel(channel, 0, force_rescrape=True)
                    print("\nScraping completed!")
            case 'n':
                selected_channels = await select_channels_to_scrape()
                if selected_channels:
                    print(f"\nChecking {len(selected_channels)} channel(s) for new messages...")
                    for channel in selected_channels:
                        last_id = state['channels'][channel]['last_id']
                        await scrape_new_messages(channel, last_id)
                    print("\nNew message check completed!")
            case 'm':
                state['scrape_media'] = not state['scrape_media']
                save_channels(state['channels'])
                print(
                    f"Media scraping {'enabled' if state['scrape_media'] else 'disabled'}.")
            case 'c':
                global continuous_scraping_active, continuous_scraping_task
                if continuous_scraping_active:
                    print("Continuous scraping is already running!")
                else:
                    continuous_scraping_active = True
                    continuous_scraping_task = asyncio.create_task(continuous_scraping())
                    print("Continuous scraping started in background. Use option [X] to stop.")
            case 'x':
                if continuous_scraping_active:
                    print("\nStopping continuous scraping...")
                    continuous_scraping_active = False
                    if continuous_scraping_task:
                        continuous_scraping_task.cancel()
                        try:
                            await continuous_scraping_task
                        except asyncio.CancelledError:
                            pass
                        continuous_scraping_task = None
                else:
                    print("Continuous scraping is not running.")
            case 'e':
                await export_data()
            case 'v':
                await view_channels()
            case 'd':
                await delete_channel_data()
            case 't':
                if not state['channels']:
                    print("\nNo channels to scrape. Please add channels first.")
                    continue
                    
                print("\nSelect channels to scrape:")
                selected_channels = await select_channels_to_scrape()
                
                if not selected_channels:
                    continue
                    
                start_date, end_date = get_date_range()
                
                print(f"\nScraping {len(selected_channels)} channel(s) from {start_date} to {end_date}...")
                for channel in selected_channels:
                    await scrape_date_range(channel, start_date, end_date)
                print("\nDate range scraping completed!")
            case 'q':
                print("Quitting...")
                sys.exit()
            case 'l':
                await add_from_channel_list()
            case 'o':
                should_exit = await logout_and_clear_credentials()
                if should_exit:
                    print("Logged out successfully. Exiting...")
                    sys.exit()

            case _:
                print("Invalid option.")

async def logout_and_clear_credentials():
    try:
        await client.log_out()
        if os.path.exists(CREDENTIALS_FILE):
            os.remove(CREDENTIALS_FILE)
        if os.path.exists('session'):
            os.remove('session')
        print("\nSuccessfully logged out and cleared credentials.")
        
        while True:
            choice = input("\nWould you like to:\n[1] Log in with different credentials\n[2] Quit\n\nChoice: ").strip()
            if choice == '1':
                print("\nRestarting program with new login...")
                # Instead of trying to recreate client, we'll restart the whole program
                python = sys.executable
                os.execl(python, python, *sys.argv)
            elif choice == '2':
                return True  # Exit from manage_channels
            else:
                print("Invalid choice. Please enter 1 or 2.")
                
    except Exception as e:
        print(f"\nError during logout: {str(e)}")
        return True

async def main():
    print("\nConnecting to Telegram...")
    try:
        await client.connect()
        
        if not await client.is_user_authorized():
            if not state['api_id'] or not state['api_hash'] or not state['phone']:
                state['api_id'] = int(input("Enter your API ID: "))
                state['api_hash'] = input("Enter your API Hash: ")
                state['phone'] = input("Enter your phone number (with country code): ")
            
            try:
                print(f"\nSending authentication code to {state['phone']}")
                await client.send_code_request(phone=state['phone'])
                verification_code = input('\nEnter the code you received on Telegram: ')
                
                try:
                    await client.sign_in(phone=state['phone'], code=verification_code)
                except SessionPasswordNeededError:
                    password = input('Two-step verification enabled. Please enter your password: ')
                    await client.sign_in(password=password)
                
                # Save credentials after successful login
                session_string = StringSession.save(client.session)
                save_credentials(state['api_id'], state['api_hash'], state['phone'], session_string)
                print("\nSuccessfully authenticated and saved credentials!")
                
            except PhoneNumberInvalidError:
                print("\nThe phone number you entered is invalid. Please include the country code.")
                print("Example: +14155552671 for US number")
                state['phone'] = None
                sys.exit(1)
            except Exception as e:
                print(f"\nError during authentication: {str(e)}")
                sys.exit(1)
        
        print("\nSuccessfully connected to Telegram!")
        while True:
            await manage_channels()
            
    except Exception as e:
        print(f"\nConnection error: {str(e)}")
        # If connection fails, clear credentials and restart
        if "authorization" in str(e).lower():
            print("\nAuthorization failed. Please log in again.")
            await logout_and_clear_credentials()
            state['api_id'] = None
            state['api_hash'] = None
            state['phone'] = None
        sys.exit(1)

if __name__ == '__main__':
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\nProgram interrupted. Exiting...")
        sys.exit()
